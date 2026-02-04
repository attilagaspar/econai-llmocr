import os
import json
import argparse
from pathlib import Path
import openpyxl
from openpyxl.styles import Alignment

# Type useful labels (same as in layout_superstructure_detect.py)
TYPE_USEFUL = ["tablazatelem", "tablazatfejlec"]


def identify_table_groups(shapes):
    """Group shapes into separate tables based on super_row/super_column sequences.
    
    Tables are identified by detecting when super_row numbering restarts (goes back to 1).
    This indicates a new table region on the same page.
    """
    # Filter for TYPE_USEFUL shapes that have super_row and super_column
    useful_shapes = [
        s for s in shapes 
        if s.get("label") in TYPE_USEFUL 
        and "super_row" in s 
        and "super_column" in s
    ]
    
    if not useful_shapes:
        return []
    
    # Sort shapes by their vertical position (top to bottom)
    def get_y_min(shape):
        points = shape.get("points", [])
        if len(points) >= 2:
            return min(p[1] for p in points)
        return 0
    
    useful_shapes.sort(key=get_y_min)
    
    # Group shapes by detecting when super_row restarts
    # When we see super_row=1 again after having seen higher row numbers, it indicates a new table
    tables = []
    current_table = []
    max_row_seen = 0
    seen_header = False
    
    for shape in useful_shapes:
        super_row = shape.get("super_row", 0)
        
        # Detect table restart: if we see super_row=1 again after having already seen it and higher rows
        if super_row == 1:
            if seen_header and max_row_seen > 1:
                # This is a new table header - save the current table and start a new one
                if current_table:
                    tables.append(current_table)
                current_table = [shape]
                max_row_seen = 1
                seen_header = True
            else:
                # First header of current table
                current_table.append(shape)
                seen_header = True
                max_row_seen = max(max_row_seen, 1)
        else:
            # Continue with current table
            current_table.append(shape)
            max_row_seen = max(max_row_seen, super_row)
    
    # Don't forget the last table
    if current_table:
        tables.append(current_table)
    
    return tables


def export_table_to_xlsx(shapes, output_path, data_field):
    """Export a table to an XLSX file.
    
    Args:
        shapes: List of shape dictionaries with super_row and super_column
        output_path: Path to save the XLSX file
        data_field: Name of the field to extract from shapes (e.g., 'original_pdf_text_layer')
    """
    if not shapes:
        return False
    
    # Find the range of rows and columns
    rows = set()
    cols = set()
    for shape in shapes:
        if "super_row" in shape and "super_column" in shape:
            rows.add(shape["super_row"])
            cols.add(shape["super_column"])
    
    if not rows or not cols:
        return False
    
    min_row, max_row = min(rows), max(rows)
    min_col, max_col = min(cols), max(cols)
    
    # Create a workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Create a mapping from (row, col) to data
    table_data = {}
    for shape in shapes:
        if "super_row" in shape and "super_column" in shape:
            row = shape["super_row"]
            col = shape["super_column"]
            data = shape.get(data_field, "")
            table_data[(row, col)] = data
    
    # Write to Excel - map super_row/super_column to Excel coordinates
    for row_idx in range(min_row, max_row + 1):
        for col_idx in range(min_col, max_col + 1):
            data = table_data.get((row_idx, col_idx), "")
            # Excel uses 1-based indexing, adjust for super_row/super_column
            excel_row = row_idx - min_row + 1
            excel_col = col_idx - min_col + 1
            cell = ws.cell(row=excel_row, column=excel_col, value=data)
            # Set text wrapping and top alignment
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Auto-adjust column widths (set to reasonable default)
    for col_idx in range(1, max_col - min_col + 2):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20
    
    # Save
    wb.save(output_path)
    return True


def get_super_column_count(table_shapes):
    """Get the number of unique super_columns in a table."""
    super_columns = set()
    for shape in table_shapes:
        if "super_column" in shape:
            super_columns.add(shape["super_column"])
    return len(super_columns)


def process_json_file(json_path, output_dir, data_field, input_root, group_by_columns=True):
    """Process a single JSON file and export tables.
    
    Args:
        json_path: Path to the JSON file
        output_dir: Output directory for Excel files
        data_field: Field to extract from shapes
        input_root: Root input directory
        group_by_columns: If True, group tables by super_column count
    
    Returns:
        Number of tables exported
    """
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        shapes = data.get("shapes", [])
        
        # Identify separate tables (regions)
        table_groups = identify_table_groups(shapes)
        
        if not table_groups:
            return 0
        
        # Generate output filenames based on relative path
        rel_path = os.path.relpath(json_path, input_root)
        # Remove .json extension and replace path separators with underscores
        base_name = rel_path.replace('.json', '').replace(os.sep, '_').replace('/', '_')
        
        exported_count = 0
        
        if group_by_columns:
            # Group tables by their super_column count
            tables_by_column_count = {}
            for table_shapes in table_groups:
                col_count = get_super_column_count(table_shapes)
                if col_count not in tables_by_column_count:
                    tables_by_column_count[col_count] = []
                tables_by_column_count[col_count].append(table_shapes)
            
            # Export each group to a separate Excel file
            for col_count, tables in tables_by_column_count.items():
                output_filename = f"{base_name}_{col_count}columns.xlsx"
                output_path = os.path.join(output_dir, output_filename)
                
                # Create a workbook for all tables with this column count
                wb = openpyxl.Workbook()
                wb.remove(wb.active)  # Remove default sheet
                
                for table_idx, table_shapes in enumerate(tables, start=1):
                    # Create a new sheet for each table
                    ws = wb.create_sheet(title=f"Table{table_idx}")
                    
                    # Export table to this sheet
                    if not table_shapes:
                        continue
                    
                    # Find the range of rows and columns
                    rows = set()
                    cols = set()
                    for shape in table_shapes:
                        if "super_row" in shape and "super_column" in shape:
                            rows.add(shape["super_row"])
                            cols.add(shape["super_column"])
                    
                    if not rows or not cols:
                        continue
                    
                    min_row, max_row = min(rows), max(rows)
                    min_col, max_col = min(cols), max(cols)
                    
                    # Create a mapping from (row, col) to data
                    table_data = {}
                    for shape in table_shapes:
                        if "super_row" in shape and "super_column" in shape:
                            row = shape["super_row"]
                            col = shape["super_column"]
                            data = shape.get(data_field, "")
                            table_data[(row, col)] = data
                    
                    # Write to Excel sheet
                    for row_idx in range(min_row, max_row + 1):
                        for col_idx in range(min_col, max_col + 1):
                            data = table_data.get((row_idx, col_idx), "")
                            excel_row = row_idx - min_row + 1
                            excel_col = col_idx - min_col + 1
                            cell = ws.cell(row=excel_row, column=excel_col, value=data)
                            cell.alignment = Alignment(wrap_text=True, vertical='top')
                    
                    # Auto-adjust column widths
                    for col_idx in range(1, max_col - min_col + 2):
                        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20
                
                # Save workbook
                wb.save(output_path)
                print(f"Exported: {output_filename} ({len(tables)} tables with {col_count} columns)")
                exported_count += len(tables)
        else:
            # Original behavior: export each table separately
            for table_idx, table_shapes in enumerate(table_groups, start=1):
                output_filename = f"{base_name}_table{table_idx}.xlsx"
                output_path = os.path.join(output_dir, output_filename)
                
                if export_table_to_xlsx(table_shapes, output_path, data_field):
                    print(f"Exported: {output_filename} ({len(table_shapes)} shapes)")
                    exported_count += 1
        
        return exported_count
        
    except Exception as e:
        print(f"Error processing {json_path}: {e}")
        return 0


def main():
    parser = argparse.ArgumentParser(
        description='Export tables from labelme JSONs with super_row/super_column to XLSX files'
    )
    parser.add_argument('input_folder', help='Input folder to search recursively for JSON files')
    parser.add_argument('output_folder', help='Output folder for XLSX files')
    parser.add_argument('--data-field', default='original_pdf_text_layer',
                        help='JSON field to export as cell content (default: original_pdf_text_layer)')
    parser.add_argument('--group-by-columns', action='store_true', default=True,
                        help='Group tables by super_column count (default: True)')
    parser.add_argument('--no-group-by-columns', dest='group_by_columns', action='store_false',
                        help='Disable grouping by super_column count')
    
    args = parser.parse_args()
    
    input_folder = args.input_folder
    output_folder = args.output_folder
    data_field = args.data_field
    group_by_columns = args.group_by_columns
    
    # Create output folder if it doesn't exist
    os.makedirs(output_folder, exist_ok=True)
    
    # Find all JSON files recursively
    json_files = []
    for root, dirs, files in os.walk(input_folder):
        for file in files:
            if file.endswith('.json'):
                json_files.append(os.path.join(root, file))
    
    print(f"Found {len(json_files)} JSON files")
    print(f"Exporting '{data_field}' field to XLSX files...")
    print(f"Group by columns: {group_by_columns}")
    print()
    
    total_tables = 0
    processed_files = 0
    
    for json_file in json_files:
        tables_exported = process_json_file(json_file, output_folder, data_field, input_folder, group_by_columns)
        if tables_exported > 0:
            processed_files += 1
            total_tables += tables_exported
    
    print()
    print(f"=== EXPORT COMPLETE ===")
    print(f"Processed {processed_files} JSON files with tables")
    print(f"Exported {total_tables} XLSX files to {output_folder}")


if __name__ == "__main__":
    main()
