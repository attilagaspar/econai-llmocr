import sys
import os
import json
import pandas as pd
import re
from openpyxl.styles import PatternFill

DOUBLE_PAGE = True


def clean_excel_string(s):
    # Remove illegal characters for Excel (openpyxl)
    if not isinstance(s, str):
        return s
    # Remove all control characters except tab (\t), newline (\n), and carriage return (\r)
    return re.sub(r"[\x00-\x08\x0b-\x0c\x0e-\x1f]", "", s)

def get_final_value(shape):
    # Priority: human_output.human_corrected_text > openai_output.response > tesseract_output.ocr_text (lines)
    if "human_output" in shape and "human_corrected_text" in shape["human_output"]:
        return shape["human_output"]["human_corrected_text"], "human"
    #elif shape.get("label") in ("text_cell", "column_header") and "openai_output" in shape and "response" in shape["openai_output"]:
    #    return shape["openai_output"]["response"], "llm"
    #elif shape.get("label") == "numerical_cell" and "tesseract_output" in shape and "ocr_text" in shape["tesseract_output"]:
    #    return shape["tesseract_output"]["ocr_text"], "ocr"
    else:
        return "", "none"

def natural_key(s):
    """Sort helper: splits string into list of ints and strings for natural sorting."""
    return [int(text) if text.isdigit() else text.lower() for text in re.split(r'(\d+)', s)]


def natural_path_key(path):
    # Natural sort for full relative path
    parts = re.split(r'([\\/])', path)
    return [natural_key(part) if part not in ('/', '\\') else part for part in parts]

def get_shape_width(shape):
    points = shape.get("points", [])
    if len(points) == 2:
        x1, _ = points[0]
        x2, _ = points[1]
        return abs(x2 - x1)
    return float("inf")

def process_json(json_path, column_filter=None, human_only=False):
    print(f"Processing JSON: {json_path}")
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    shapes = [s for s in data.get("shapes", []) if get_shape_width(s) < 10000]
    print(f"  Found {len(shapes)} shapes narrower than 500px")
    # Group shapes by super_row
    rows = {}
    sources = {}
    for shape in shapes:
        row = shape.get("super_row")
        col = shape.get("super_column")
        if row is None or col is None:
            print(f"    Skipping shape without super_row or super_column")
            continue
        
        # Apply column filter if specified
        if column_filter is not None and col not in column_filter:
            continue
        
        # Apply human-only filter if specified
        if human_only and not ("human_output" in shape and "human_corrected_text" in shape["human_output"]):
            continue
            
        val, source = get_final_value(shape)
        if row not in rows:
            rows[row] = {}
            sources[row] = {}
        rows[row][col] = val
        sources[row][col] = source

    # For each super_row, split each cell into lines and pad to max block height
    row_blocks = []
    source_blocks = []
    row_counter = 1  # within-page row number
    for row in sorted(rows.keys()):
        cols = rows[row]
        srcs = sources[row]
        # Split each cell into lines
        cell_lines = {col: str(cols[col]).split('\n') for col in cols}
        cell_sources = {col: [srcs[col]] * len(cell_lines[col]) for col in cols}
        max_lines = max(len(lines) for lines in cell_lines.values()) if cell_lines else 1
        print(f"  Row {row}: max block height = {max_lines}")
        # Pad each cell to max_lines
        for col in cell_lines:
            if len(cell_lines[col]) < max_lines:
                print(f"    Padding column {col} in row {row} from {len(cell_lines[col])} to {max_lines} lines")
                pad_len = max_lines - len(cell_lines[col])
                cell_lines[col] += [""] * pad_len
                cell_sources[col] += [srcs[col]] * pad_len
        # For each line, build a row for excel (each line is a new row)
        for i in range(max_lines):
            excel_row = {
                "source_json_path": json_path,
                "source_directory": os.path.dirname(json_path),
                "source_json": os.path.basename(json_path),
                "within_json_row": row_counter,
                "_super_row": row
            }
            source_row = {"_super_row": row}
            human_flag = 0
            for col in cell_lines:
                excel_row[col] = cell_lines[col][i]
                source_row[col] = cell_sources[col][i]
                if cell_sources[col][i] == "human":
                    human_flag = 1
            excel_row["any_human_output"] = human_flag
            row_blocks.append(excel_row)
            source_blocks.append(source_row)
            row_counter += 1
    print(f"  Generated {len(row_blocks)} excel rows for this JSON")
    return row_blocks, source_blocks

def main(input_folder, output_excel, column_filter=None, human_only=False, double_page=False):
    print(f"Scanning folder recursively: {input_folder}")
    if column_filter:
        print(f"Filtering for super_columns: {column_filter}")
    if human_only:
        print(f"Filtering for human output only")
    if double_page:
        print(f"Using double-page mode: even pages extend odd pages horizontally")
    
    json_files = []
    for root, _, files in os.walk(input_folder):
        for fname in files:
            if fname.lower().endswith(".json"):
                json_files.append(os.path.join(root, fname))
    print(f"Found {len(json_files)} JSON files.")
    # Sort JSONs in natural order (page_5, page_11, etc.)
    json_files = sorted(json_files, key=natural_path_key)
    
    if double_page:
        return process_double_page_mode(json_files, output_excel, column_filter, human_only)
    else:
        return process_standard_mode(json_files, output_excel, column_filter, human_only)

def process_standard_mode(json_files, output_excel, column_filter, human_only):
    all_rows = []
    all_sources = []
    for json_path in json_files:
        row_blocks, source_blocks = process_json(json_path, column_filter, human_only)
        all_rows.extend(row_blocks)
        all_sources.extend(source_blocks)
        print(f"Appended {len(row_blocks)} rows from {os.path.basename(json_path)}")
    # Find all super_columns used
    all_cols = set()
    for row in all_rows:
        all_cols.update([c for c in row if c not in ("_super_row", "source_json_path", "source_directory", "source_json", "within_json_row", "any_human_output")])
    sorted_cols = sorted(all_cols, key=lambda x: int(x))
    print(f"Detected columns: {sorted_cols}")
    return finalize_excel_export(all_rows, all_sources, sorted_cols, output_excel)

def process_double_page_mode(json_files, output_excel, column_filter, human_only):
    final_rows = []
    final_sources = []
    current_row_offset = 0
    
    # Process JSONs in pairs (odd + even)
    for i in range(0, len(json_files), 2):
        odd_json = json_files[i]
        even_json = json_files[i + 1] if i + 1 < len(json_files) else None
        
        # Process odd page (left page)
        odd_rows, odd_sources = process_json(odd_json, column_filter, human_only)
        print(f"Processed odd page {os.path.basename(odd_json)}: {len(odd_rows)} rows")
        
        # Process even page (right page) if it exists
        if even_json:
            even_rows, even_sources = process_json(even_json, column_filter, human_only)
            print(f"Processed even page {os.path.basename(even_json)}: {len(even_rows)} rows")
        else:
            even_rows, even_sources = [], []
            print(f"No even page to pair with {os.path.basename(odd_json)}")
        
        # Merge the two pages horizontally
        max_rows = max(len(odd_rows), len(even_rows))
        
        for row_idx in range(max_rows):
            merged_row = {
                "source_json_path": f"{odd_json}" + (f" + {even_json}" if even_json else ""),
                "source_directory": os.path.dirname(odd_json),
                "source_json": f"{os.path.basename(odd_json)}" + (f" + {os.path.basename(even_json)}" if even_json else ""),
                "within_json_row": row_idx + 1,
                "_super_row": current_row_offset + row_idx + 1,
                "any_human_output": 0
            }
            merged_source = {"_super_row": current_row_offset + row_idx + 1}
            
            # Add odd page data
            if row_idx < len(odd_rows):
                odd_row = odd_rows[row_idx]
                odd_source = odd_sources[row_idx]
                for key, value in odd_row.items():
                    if key not in ("source_json_path", "source_directory", "source_json", "within_json_row", "_super_row", "any_human_output"):
                        merged_row[key] = value
                        merged_source[key] = odd_source.get(key, "none")
                        if odd_source.get(key) == "human":
                            merged_row["any_human_output"] = 1
            
            # Add even page data with column offset
            if row_idx < len(even_rows) and even_json:
                even_row = even_rows[row_idx]
                even_source = even_sources[row_idx]
                
                # Find the maximum column number from odd page to determine offset
                odd_cols = []
                for k in merged_row.keys():
                    if isinstance(k, str) and k.isdigit():
                        odd_cols.append(int(k))
                    elif isinstance(k, int):
                        odd_cols.append(k)
                col_offset = max(odd_cols) + 1 if odd_cols else 1
                
                for key, value in even_row.items():
                    # Check if key represents a column (either string digit or integer)
                    is_column = False
                    if isinstance(key, str) and key.isdigit():
                        is_column = True
                    elif isinstance(key, int):
                        is_column = True
                    
                    if key not in ("source_json_path", "source_directory", "source_json", "within_json_row", "_super_row", "any_human_output") and is_column:
                        key_num = int(key) if isinstance(key, str) else key
                        new_col = str(key_num + col_offset - 1)  # Adjust column numbering
                        merged_row[new_col] = value
                        merged_source[new_col] = even_source.get(key, "none")
                        if even_source.get(key) == "human":
                            merged_row["any_human_output"] = 1
            
            final_rows.append(merged_row)
            final_sources.append(merged_source)
        
        current_row_offset += max_rows
        print(f"Merged pages into {max_rows} rows, next row offset: {current_row_offset}")
    
    # Find all super_columns used
    all_cols = set()
    for row in final_rows:
        all_cols.update([c for c in row if c not in ("_super_row", "source_json_path", "source_directory", "source_json", "within_json_row", "any_human_output")])
    sorted_cols = sorted(all_cols, key=lambda x: int(x))
    print(f"Detected columns after double-page merge: {sorted_cols}")
    return finalize_excel_export(final_rows, final_sources, sorted_cols, output_excel)

def finalize_excel_export(all_rows, all_sources, sorted_cols, output_excel):

    # sanitize for excel export
    for row in all_rows:
        for k in row:
            row[k] = clean_excel_string(row[k])
    df = pd.DataFrame(all_rows)
    df = df[["source_json_path", "source_directory", "source_json", "within_json_row", "any_human_output", "_super_row"] + sorted_cols]

    # Write to Excel with coloring
    print(f"Writing {len(df)} rows to Excel: {output_excel}")
    with pd.ExcelWriter(output_excel, engine="openpyxl") as writer:
        df.to_excel(writer, index=False)
        workbook = writer.book
        worksheet = writer.sheets["Sheet1"]
        # Apply light blue fill to cells from human_output
        light_blue = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type="solid")
        for row_idx, source_row in enumerate(all_sources, start=2):  # Excel rows start at 2 (header is row 1)
            for col_idx, col in enumerate(sorted_cols, start=7):     # Excel cols start at 7 (source_json_path=1, source_directory=2, source_json=3, within_json_row=4, any_human_output=5, _super_row=6)
                if source_row.get(col) == "human":
                    worksheet.cell(row=row_idx, column=col_idx).fill = light_blue
    print(f"Exported to {output_excel}")
    return df

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python json_join_excel_export.py input_json_folder output_excel_file.xlsx [-h] [-d] [super_column1 super_column2 ...]")
        print("  -h: Export only elements with human_output and human_corrected_text")
        print("  -d: Double-page mode: even pages extend odd pages horizontally")
        print("Example: python json_join_excel_export.py input_folder export.xlsx 2 3 9")
        print("Example: python json_join_excel_export.py input_folder export.xlsx -h")
        print("Example: python json_join_excel_export.py input_folder export.xlsx -d")
        print("Example: python json_join_excel_export.py input_folder export.xlsx -h -d 2 3 9")
        sys.exit(1)
    
    input_folder = sys.argv[1]
    output_excel = sys.argv[2]
    
    # Parse arguments
    args = sys.argv[3:]
    human_only = False
    double_page = False
    column_filter = None
    
    # Check for -h flag
    if "-h" in args:
        human_only = True
        args.remove("-h")
    
    # Check for -d flag
    if "-d" in args:
        double_page = True
        args.remove("-d")
    
    # Parse remaining arguments as column filters
    if args:
        try:
            column_filter = [int(arg) for arg in args]
        except ValueError:
            print("Error: Column filter arguments must be integers")
            sys.exit(1)
    
    main(input_folder, output_excel, column_filter, human_only, double_page)
    print("Done processing JSON files and exporting to Excel.")