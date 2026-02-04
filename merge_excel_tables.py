#!/usr/bin/env python3
"""
Merge multiple Excel files into a single master Excel file.
Appends all tables one after another with a source filename column.
"""

import argparse
import os
import re
from pathlib import Path
from openpyxl import Workbook, load_workbook


def natural_sort_key(filename):
    """
    Natural sort key for filenames like page_1, page_2, page_10, page_15_table2, etc.
    Converts numeric parts to integers for proper sorting.
    """
    def atoi(text):
        return int(text) if text.isdigit() else text.lower()
    
    return [atoi(c) for c in re.split(r'(\d+)', filename)]


def extract_column_count(filename):
    """
    Extract column count from filename ending like '_6columns.xlsx' or '_7columns.xlsx'.
    Returns the column count as an integer, or None if no column count pattern found.
    """
    match = re.search(r'_(\d+)columns\.xlsx$', filename, re.IGNORECASE)
    if match:
        return int(match.group(1))
    return None


def merge_excel_files(input_folder, output_file, group_by_columns=True):
    """
    Merge all XLSX files from input_folder into a single output_file (or multiple files if grouped).
    Adds a 'Source File' column to track which file each row came from.
    
    Args:
        input_folder: Folder containing XLSX files to merge
        output_file: Output file path (base name if grouping by columns)
        group_by_columns: If True, group files by column count and create separate output files
    """
    input_path = Path(input_folder)
    
    if not input_path.exists() or not input_path.is_dir():
        print(f"Error: Input folder '{input_folder}' does not exist or is not a directory.")
        return
    
    # Find all XLSX files
    all_xlsx_files = sorted(input_path.glob("*.xlsx"), key=lambda p: natural_sort_key(p.stem))
    
    if not all_xlsx_files:
        print(f"No XLSX files found in '{input_folder}'")
        return
    
    print(f"Found {len(all_xlsx_files)} XLSX files")
    
    # Group files by column count if enabled
    if group_by_columns:
        files_by_columns = {}
        files_without_pattern = []
        
        for xlsx_file in all_xlsx_files:
            col_count = extract_column_count(xlsx_file.name)
            if col_count is not None:
                if col_count not in files_by_columns:
                    files_by_columns[col_count] = []
                files_by_columns[col_count].append(xlsx_file)
            else:
                files_without_pattern.append(xlsx_file)
        
        if files_without_pattern:
            print(f"  {len(files_without_pattern)} files without column pattern will be merged separately")
        
        print(f"  Found {len(files_by_columns)} different column counts")
        for col_count in sorted(files_by_columns.keys()):
            print(f"    {col_count} columns: {len(files_by_columns[col_count])} files")
        print()
        
        # Process each column count group separately
        for col_count in sorted(files_by_columns.keys()):
            xlsx_files = files_by_columns[col_count]
            # Create output filename with column count
            output_path = Path(output_file)
            output_file_with_columns = str(output_path.with_stem(f"{output_path.stem}_{col_count}columns"))
            
            print(f"\n=== Processing files with {col_count} columns ===")
            _merge_files_to_single_output(xlsx_files, output_file_with_columns)
        
        # Process files without column pattern if any
        if files_without_pattern:
            print(f"\n=== Processing files without column pattern ===")
            output_path = Path(output_file)
            output_file_no_pattern = str(output_path.with_stem(f"{output_path.stem}_other"))
            _merge_files_to_single_output(files_without_pattern, output_file_no_pattern)
        
        return
    
    # Original behavior: merge all files into single output
    print(f"Merging all {len(all_xlsx_files)} files into single output")
    _merge_files_to_single_output(all_xlsx_files, output_file)


def _merge_files_to_single_output(xlsx_files, output_file):
    """
    Internal function to merge a list of XLSX files into a single output file.
    """
    print(f"Merging {len(xlsx_files)} files to: {output_file}")
    
    # Create master workbook
    master_wb = Workbook()
    master_ws = master_wb.active
    master_ws.title = "Merged Tables"
    
    current_row = 1
    first_file = True
    
    for xlsx_file in xlsx_files:
        print(f"  Processing: {xlsx_file.name}")
        
        try:
            wb = load_workbook(xlsx_file, data_only=True)
            ws = wb.active
            
            # Get max row and column
            max_row = ws.max_row
            max_col = ws.max_column
            
            if max_row == 0 or max_col == 0:
                print(f"    Skipping empty file: {xlsx_file.name}")
                continue
            
            # Add header with "Source File" column on first file only
            if first_file:
                # Copy header row
                for col_idx in range(1, max_col + 1):
                    cell_value = ws.cell(row=1, column=col_idx).value
                    master_ws.cell(row=current_row, column=col_idx, value=cell_value)
                
                # Add "Source File" header
                master_ws.cell(row=current_row, column=max_col + 1, value="Source File")
                current_row += 1
                first_file = False
            
            # Copy data rows (skip header row for all files)
            for row_idx in range(2, max_row + 1):
                for col_idx in range(1, max_col + 1):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    master_ws.cell(row=current_row, column=col_idx, value=cell_value)
                
                # Add source filename
                master_ws.cell(row=current_row, column=max_col + 1, value=xlsx_file.name)
                current_row += 1
            
            wb.close()
            print(f"    Added {max_row - 1} rows")
            
        except Exception as e:
            print(f"    Error processing {xlsx_file.name}: {e}")
            continue
    
    # Save master workbook
    try:
        master_wb.save(output_file)
        print(f"  === MERGE COMPLETE ===")
        print(f"  Merged {len(xlsx_files)} files")
        print(f"  Total rows (including header): {current_row - 1}")
        print(f"  Output saved to: {output_file}")
    except Exception as e:
        print(f"  Error saving output file: {e}")


def main():
    parser = argparse.ArgumentParser(
        description="Merge multiple Excel tables into a single master Excel file"
    )
    parser.add_argument(
        "input_folder",
        help="Folder containing XLSX files to merge"
    )
    parser.add_argument(
        "output_file",
        help="Output master Excel file path (base name if grouping by columns)"
    )
    parser.add_argument(
        "--group-by-columns",
        action="store_true",
        default=True,
        help="Group files by column count pattern (default: True)"
    )
    parser.add_argument(
        "--no-group-by-columns",
        dest="group_by_columns",
        action="store_false",
        help="Disable grouping by column count"
    )
    
    args = parser.parse_args()
    
    merge_excel_files(args.input_folder, args.output_file, args.group_by_columns)


if __name__ == "__main__":
    main()
